import os
import shutil
from datetime import datetime
from PIL import Image
from PIL.ExifTags import TAGS
from pymediainfo import MediaInfo
from concurrent.futures import ThreadPoolExecutor
from threading import Lock
import json
from openpyxl import Workbook, load_workbook

# -------------------------
# Configuration
# -------------------------
source_folder = r"/path/to/source_folder"
destination_folder = r"/path/to/destination_folder"

skip_extensions = {".json"}
image_ext = {".jpg", ".jpeg", ".png", ".heic", ".webp"}
video_ext = {".mp4", ".mov", ".3gp", ".mkv", ".ts"}

os.makedirs(destination_folder, exist_ok=True)

counter_lock = Lock()
processed = 0
failed_files = []

log_file_path = os.path.join(destination_folder, "file_moves.xlsx")
log_lock = Lock()

# -------------------------
# Initialize Excel log
# -------------------------
if not os.path.exists(log_file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "File Moves"
    ws.append(["Original Path", "New Path", "Method Used", "File Name Without Ext", "Only Ext"])
    wb.save(log_file_path)

# -------------------------
# Build JSON metadata index
# -------------------------
print("Indexing all JSON files...")
json_index = {}  # base filename without extension -> full JSON path
for root, dirs, files in os.walk(source_folder):
    for f in files:
        if f.lower().endswith(".json"):
            key = os.path.splitext(f)[0]  # base filename
            json_index[key] = os.path.join(root, f)
print(f"Found {len(json_index)} JSON metadata files.")

# -------------------------
# JSON fallback
# -------------------------
def get_json_timestamp(path):
    filename = os.path.splitext(os.path.basename(path))[0]
    json_path = json_index.get(filename)
    if json_path:
        try:
            with open(json_path, "r", encoding="utf-8") as jf:
                data = json.load(jf)
                ts = int(data.get("photoTakenTime", {}).get("timestamp", 0))
                if ts:
                    return datetime.fromtimestamp(ts), "JSON"
        except:
            pass
    return None, None

# -------------------------
# Image date extraction
# -------------------------
def get_photo_date(path):
    date_from_json, method = get_json_timestamp(path)
    if date_from_json:
        return date_from_json, method
    try:
        image = Image.open(path)
        exif = image._getexif()
        if exif:
            exif_data = {TAGS.get(tag, tag): value for tag, value in exif.items()}
            for key in ["DateTimeOriginal", "DateTimeDigitized"]:
                if key in exif_data:
                    return datetime.strptime(exif_data[key], "%Y:%m:%d %H:%M:%S"), "EXIF"
    except:
        pass
    return None, None

# -------------------------
# Video date extraction
# -------------------------
def get_video_date(path):
    date_from_json, method = get_json_timestamp(path)
    if date_from_json:
        return date_from_json, method
    try:
        media_info = MediaInfo.parse(path)
        for track in media_info.tracks:
            if track.track_type == "General":
                value = getattr(track, "encoded_date", None) or getattr(track, "tagged_date", None)
                if value:
                    value = value.replace("UTC", "").strip()
                    try:
                        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S"), "MediaInfo"
                    except:
                        continue
    except:
        pass
    return None, None

# -------------------------
# File processing
# -------------------------
def process_file(file_path):
    global processed
    if os.path.abspath(file_path).startswith(os.path.abspath(destination_folder)):
        return

    file = os.path.basename(file_path)
    ext = os.path.splitext(file)[1].lower()
    if ext in skip_extensions:
        return

    date_taken = None
    media_type = None
    method_used = "Unknown"

    try:
        if ext in image_ext:
            media_type = "Pictures"
            date_taken, method_used = get_photo_date(file_path)
        elif ext in video_ext:
            media_type = "Videos"
            date_taken, method_used = get_video_date(file_path)
        else:
            return

        if date_taken:
            year_folder = date_taken.strftime("%Y")
            month_folder = date_taken.strftime("%m")
            target_folder = os.path.join(destination_folder, year_folder, month_folder, media_type)
        else:
            target_folder = os.path.join(destination_folder, "No-Date-Taken", media_type)

        os.makedirs(target_folder, exist_ok=True)

        destination_path = os.path.join(target_folder, file)
        counter = 1
        base_name, extension = os.path.splitext(file)
        while os.path.exists(destination_path):
            destination_path = os.path.join(target_folder, f"{base_name}_{counter}{extension}")
            counter += 1

        shutil.move(file_path, destination_path)

        # -------------------------
        # Log to Excel
        # -------------------------
        with log_lock:
            wb = load_workbook(log_file_path)
            ws = wb.active
            ws.append([
                file_path,
                destination_path,
                method_used,
                os.path.splitext(os.path.basename(destination_path))[0],  # File Name Without Ext
                os.path.splitext(os.path.basename(destination_path))[1]   # Only Extension
            ])
            wb.save(log_file_path)

    except Exception as e:
        with counter_lock:
            failed_files.append((file_path, str(e)))

    with counter_lock:
        processed += 1
        if processed % 50 == 0:
            print(f"Processed {processed} files...")

# -------------------------
# Gather all files
# -------------------------
all_files = []
for root, dirs, files in os.walk(source_folder):
    if os.path.abspath(root).startswith(os.path.abspath(destination_folder)):
        continue
    for file in files:
        all_files.append(os.path.join(root, file))

print(f"Total files found: {len(all_files)}")

# -------------------------
# Process with multithreading
# -------------------------
with ThreadPoolExecutor(max_workers=8) as executor:
    executor.map(process_file, all_files)

print(f"Finished processing {processed} files.")

if failed_files:
    print("Some files failed to move:")
    for f, e in failed_files:
        print(f"{f} --> {e}")

print(f"All moves are logged in: {log_file_path}")
input("Press Enter to exit...")
